#!/usr/bin/env python3
"""
generate_masterlist.py
======================
Konvertiert "Inventory value by item number, planned status, and age.xlsx"
in masterlist.json für die Inventur-PWA.

Aufruf:
    python generate_masterlist.py
    python generate_masterlist.py --input "pfad/zur/datei.xlsx" --out data/masterlist.json

Wöchentlich ausführen wenn eine neue Masterlist-XLSX eintrifft.
"""

import sys
import os
import json
import argparse
import subprocess
import datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Fehlende Abhängigkeit: openpyxl")
    print("  → pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Standardpfade (relativ zu diesem Script)
# ---------------------------------------------------------------------------
DEFAULT_INPUT = Path(__file__).parent.parent / "PART_MGMT" / "input" / \
                "Inventory value by item number, planned status, and age.xlsx"
DEFAULT_OUT   = Path(__file__).parent / "data" / "masterlist.json"

# ---------------------------------------------------------------------------
# Spaltenindizes (0-basiert, aus der XLSX-Struktur)
# ---------------------------------------------------------------------------
COL_SUB   = 3   # SUB INVENTORY           z.B. "CHU.8678"
COL_ENG   = 4   # ENGINEER/LOCATION NAME  z.B. "8678 Edubook CSP341555xxx"
COL_ITEM  = 7   # ITEM NUMBER             z.B. "0001070128700"
COL_NAME  = 8   # ITEM NAME               z.B. "BELT-TIMING-1830-2MR09"
COL_MIN   = 17  # MIN QTY
COL_QOH   = 19  # TOTAL ON HAND QTY
COL_VAL   = 21  # TOTAL STOCK VALUE
COL_AGE   = 23  # AGE RANK                z.B. "< 30 DAYS", "Excluded"


def parse_age_rank(raw) -> int:
    """
    Konvertiert AGE RANK String zu int:
      -1 = Excluded / unbekannt
       0 = < 30 days
       1 = 30 - 60 days
       2 = 61 - 90 days
       3 = 91 - 180 days
       4 = 181 - 360 days
       5 = > 360 days
    """
    if raw is None:
        return -1
    s = str(raw).strip().upper()
    if not s or 'EXCLUDED' in s:
        return -1
    if '< 30' in s or s.startswith('<30'):
        return 0
    if '30' in s and '60' in s:
        return 1
    if '61' in s and '90' in s:
        return 2
    if '91' in s and '180' in s:
        return 3
    if '181' in s and '360' in s:
        return 4
    if '> 360' in s or '>360' in s:
        return 5
    return -1


def clean_location_name(raw: str) -> str:
    """
    Kürzt den Engineer/Location-Namen auf etwas Lesbares.
    "8678 Edubook CSP341555xxx"  →  "Edubook"
    "Baudat"                     →  "Baudat"
    """
    # Entferne CSP-Nummer am Ende
    name = raw.split('CSP')[0].strip()
    # Entferne führende 4-5stellige Nummer + Leerzeichen
    import re
    name = re.sub(r'^\d{4,5}\s+', '', name).strip()
    return name or raw


def main():
    parser = argparse.ArgumentParser(description='Masterlist XLSX → JSON für Inventur-PWA')
    parser.add_argument('--input', default=str(DEFAULT_INPUT),
                        help='Pfad zur Masterlist-XLSX')
    parser.add_argument('--out', default=str(DEFAULT_OUT),
                        help='Ausgabepfad für masterlist.json')
    parser.add_argument('--no-push', action='store_true',
                        help='Kein automatischer Git-Commit + Push')
    args = parser.parse_args()

    input_path = Path(args.input)
    out_path   = Path(args.out)

    if not input_path.exists():
        print(f"❌ Datei nicht gefunden: {input_path}")
        print("   Bitte --input Pfad angeben oder Datei in den Standardpfad legen.")
        sys.exit(1)

    print(f"Lese: {input_path.name} ...", flush=True)

    wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)
    ws = wb.active

    # _stand: Datum der letzten Änderung der Quelldatei
    mtime = os.path.getmtime(str(input_path))
    stand_date = datetime.datetime.fromtimestamp(mtime).date().isoformat()
    print(f"  Stand (Quelldatei): {stand_date}")

    # Spaltenstruktur validieren
    header = list(next(ws.iter_rows(values_only=True)))
    expected = {
        COL_SUB:  'SUB INVENTORY',
        COL_ITEM: 'ITEM NUMBER',
        COL_MIN:  'MIN QTY',
        COL_VAL:  'TOTAL STOCK VALUE',
        COL_AGE:  'AGE RANK',
    }
    for col, name in expected.items():
        actual = str(header[col]).strip() if col < len(header) else '?'
        if actual != name:
            print(f"⚠️  Spalte [{col}] erwartet '{name}', gefunden '{actual}'")
            print("   Bitte Spaltenindizes in generate_masterlist.py anpassen.")

    # Daten einlesen
    item_map   = defaultdict(dict)   # item_nr → {sub: {min, qoh}}
    item_names = {}                  # item_nr → name
    eng_names  = {}                  # sub → short location name
    row_count  = 0

    for row in ws.iter_rows(values_only=True, min_row=2):
        if len(row) <= COL_QOH:
            continue

        sub = str(row[COL_SUB]).strip() if row[COL_SUB] else ''
        if not sub or sub == 'None':
            continue

        # Item-Nummer: führende Nullen entfernen
        raw_item = str(row[COL_ITEM]).strip() if row[COL_ITEM] else ''
        item     = raw_item.lstrip('0') or raw_item
        if not item or item == 'None':
            continue

        name    = str(row[COL_NAME]).strip() if row[COL_NAME] else ''
        eng     = str(row[COL_ENG]).strip()  if row[COL_ENG]  else ''
        mn      = int(row[COL_MIN]) if isinstance(row[COL_MIN], (int, float)) else 0
        qoh     = int(row[COL_QOH]) if isinstance(row[COL_QOH], (int, float)) else 0
        val_raw = row[COL_VAL] if len(row) > COL_VAL else None
        age_raw = row[COL_AGE] if len(row) > COL_AGE else None
        val     = round(float(val_raw), 2) if isinstance(val_raw, (int, float)) else 0.0
        age     = parse_age_rank(age_raw)

        item_map[item][sub] = {'min': mn, 'qoh': qoh, 'a': age, 'v': val}

        if item not in item_names:
            item_names[item] = name
        if sub not in eng_names:
            eng_names[sub] = clean_location_name(eng)

        row_count += 1

    wb.close()

    print(f"  {row_count:,} Datenzeilen | {len(eng_names)} Lager | {len(item_map)} unique Artikel")

    # JSON aufbauen
    # Kompaktes Format: _w = warehouses dict, i = items dict
    data = {
        '_generated': datetime.date.today().isoformat(),
        '_stand':     stand_date,  # Datum der Quelldatei (mtime)
        '_w': eng_names,   # { "CHU.8678": "Edubook", ... }
        'i': {
            item: {
                'n': item_names[item],          # name
                'l': locs                        # locations: { "CHU.8678": {min, qoh} }
            }
            for item, locs in item_map.items()
        }
    }

    # Ausgabe schreiben
    out_path.parent.mkdir(parents=True, exist_ok=True)
    json_str = json.dumps(data, ensure_ascii=False, separators=(',', ':'))
    out_path.write_text(json_str, encoding='utf-8')

    size_kb = len(json_str.encode()) / 1024
    print(f"OK  Gespeichert: {out_path}  ({size_kb:.0f} KB)")

    if not args.no_push:
        git_push(out_path, data['_generated'])


def git_push(json_path: Path, date_str: str):
    """Commit masterlist.json und push zu GitHub."""
    repo_dir = json_path.parent.parent  # inventur-pwa/

    def run(cmd):
        result = subprocess.run(cmd, cwd=str(repo_dir), capture_output=True, text=True)
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip())
        return result.stdout.strip()

    print()
    try:
        # Nur pushen wenn sich die Datei tatsächlich geändert hat
        changed = run(['git', 'diff', '--name-only', str(json_path.relative_to(repo_dir))])
        if not changed:
            # Auch untracked prüfen
            status = run(['git', 'status', '--porcelain', str(json_path.relative_to(repo_dir))])
            if not status:
                print("Git: keine Aenderungen, kein Push noetig.")
                return

        run(['git', 'add', str(json_path.relative_to(repo_dir))])
        run(['git', 'commit', '-m', f'Masterlist Update {date_str}'])
        run(['git', 'push'])
        print("Git: committed + pushed -> GitHub Pages wird aktualisiert.")
    except RuntimeError as e:
        print(f"Git-Fehler: {e}")
        print("Manuell pushen: git add data/masterlist.json && git commit -m 'Update' && git push")


if __name__ == '__main__':
    main()
